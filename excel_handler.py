"""
Работа с Excel файлами
"""
from openpyxl import load_workbook, Workbook
from datetime import datetime
import os
from config import *

class ExcelHandler:
    """Класс для работы с Excel файлом"""
    
    def __init__(self, filename=EXCEL_FILE):
        self.filename = filename
        self.wb = None
        self.ws = None
    
    def load(self):
        """Загружает Excel файл"""
        if not os.path.exists(self.filename):
            print(f"❌ Файл {self.filename} не найден!")
            return False
        
        try:
            self.wb = load_workbook(self.filename)
            self.ws = self.wb[SHEET_NAME] if SHEET_NAME in self.wb.sheetnames else self.wb.active
            return True
        except Exception as e:
            print(f"❌ Ошибка при загрузке Excel: {e}")
            return False
    
    def save(self):
        """Сохраняет изменения"""
        try:
            self.wb.save(self.filename)
            return True
        except Exception as e:
            print(f"❌ Ошибка при сохранении Excel: {e}")
            return False
    
    def get_pending_requests(self):
        """Получает список невыполненных запросов"""
        pending = []
        
        for row in range(2, self.ws.max_row + 1):
            request = self.ws.cell(row, COL_REQUEST).value
            status = self.ws.cell(row, COL_STATUS).value
            
            # Берем запросы без статуса или с ошибкой
            if request and (not status or status in [STATUS_ERROR, STATUS_IN_PROGRESS, 
                                                     STATUS_RATE_LIMIT, STATUS_NETWORK_ERROR, 
                                                     STATUS_TIMEOUT]):
                
                # Читаем дополнительные параметры
                project = self._get_cell_value(row, COL_PROJECT)
                model = self._get_cell_value(row, COL_MODEL)
                chat_mode = self._get_cell_value(row, COL_CHAT_MODE)
                
                pending.append({
                    'row': row,
                    'request': request.strip(),
                    'project': project,
                    'model': model,
                    'chat_mode': chat_mode if chat_mode else CHAT_MODE_NEW
                })
        
        return pending
    
    def _get_cell_value(self, row, col):
        """Безопасно получает значение ячейки"""
        try:
            if self.ws.max_column >= col:
                value = self.ws.cell(row, col).value
                return value.strip() if value else None
            return None
        except:
            return None
    
    def update_status(self, row, status, response="", error_message=""):
        """Обновляет статус запроса"""
        self.ws.cell(row, COL_STATUS).value = status
        
        if response:
            self.ws.cell(row, COL_RESPONSE).value = response
        
        if error_message:
            self.ws.cell(row, COL_ERROR).value = error_message
        
        if status == STATUS_SUCCESS:
            self.ws.cell(row, COL_DATE).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        self.save()
    
    def get_statistics(self):
        """Возвращает статистику по запросам"""
        total = self.ws.max_row - 1  # Минус заголовок
        success = 0
        errors = 0
        pending = 0
        
        for row in range(2, self.ws.max_row + 1):
            status = self.ws.cell(row, COL_STATUS).value
            if status == STATUS_SUCCESS:
                success += 1
            elif status in [STATUS_ERROR, STATUS_RATE_LIMIT, STATUS_NETWORK_ERROR, STATUS_TIMEOUT]:
                errors += 1
            else:
                pending += 1
        
        return {
            'total': total,
            'success': success,
            'errors': errors,
            'pending': pending
        }