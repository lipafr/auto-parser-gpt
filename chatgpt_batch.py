import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from datetime import datetime
import time
import os

print("=" * 70)
print("🚀 ChatGPT Batch Parser - Пакетная обработка запросов из Excel")
print("=" * 70)

# ========== НАСТРОЙКИ ==========
EXCEL_FILE = "requests.xlsx"
SHEET_NAME = "Sheet1"

COL_REQUEST = 1  # Колонка A - Запрос
COL_RESPONSE = 2  # Колонка B - Ответ
COL_STATUS = 3  # Колонка C - Статус
COL_DATE = 4  # Колонка D - Дата

DELAY_BETWEEN_REQUESTS = 5  # Пауза между запросами (секунды)
# ================================

def load_excel():
    if not os.path.exists(EXCEL_FILE):
        print(f"❌ Файл {EXCEL_FILE} не найден!")
        print("Создайте Excel файл со столбцами: Запрос | Ответ | Статус | Дата")
        return None, None
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    return wb, ws

def save_excel(wb):
    wb.save(EXCEL_FILE)
    print(f"💾 Прогресс сохранен")

def get_pending_requests(ws):
    pending = []
    for row in range(2, ws.max_row + 1):
        request = ws.cell(row, COL_REQUEST).value
        status = ws.cell(row, COL_STATUS).value
        
        if request and (not status or status == "Ошибка" or status == "В процессе"):
            pending.append({
                'row': row,
                'request': request.strip()
            })
    
    return pending

def update_status(ws, row, status, response=""):
    ws.cell(row, COL_STATUS).value = status
    if response:
        ws.cell(row, COL_RESPONSE).value = response
    if status == "Выполнен":
        ws.cell(row, COL_DATE).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def send_to_chatgpt(driver, prompt):
    try:
        print(f"  ✏️ Ищу поле ввода...")
        
        input_box = None
        selectors = [
            (By.ID, "prompt-textarea"),
            (By.CSS_SELECTOR, "textarea[placeholder*='Message']"),
            (By.XPATH, "//textarea")
        ]
        
        for by, selector in selectors:
            try:
                input_box = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((by, selector))
                )
                if input_box:
                    print(f"  ✅ Поле найдено")
                    break
            except:
                continue
        
        if not input_box:
            print(f"  ❌ Поле ввода не найдено")
            return None
        
        print(f"  📝 Отправляю запрос...")
        input_box.click()
        time.sleep(1)
        
        # Очищаем поле
        input_box.send_keys(Keys.CONTROL + "a")
        input_box.send_keys(Keys.DELETE)
        time.sleep(0.5)
        
        # Вводим текст
        input_box.send_keys(prompt)
        time.sleep(1)
        
        # Отправляем
        input_box.send_keys(Keys.RETURN)
        
        print(f"  ⏳ Жду ответ ChatGPT...")
        time.sleep(10)
        
        # Ждем появления ответа
        try:
            WebDriverWait(driver, 90).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "[data-message-author-role='assistant']"))
            )
        except:
            print(f"  ⚠️ Ответ не появился за 90 секунд")
            return None
        
        # Ждем завершения генерации
        print(f"  ⏳ Жду завершения генерации...")
        max_wait = 120
        elapsed = 0
        
        while elapsed < max_wait:
            try:
                stop_buttons = driver.find_elements(By.XPATH, "//button[contains(text(), 'Stop')]")
                if not stop_buttons:
                    print(f"  ✅ Генерация завершена")
                    break
            except:
                break
            
            time.sleep(2)
            elapsed += 2
        
        time.sleep(3)
        
        # Читаем ответ
        print(f"  📖 Читаю ответ...")
        messages = driver.find_elements(By.CSS_SELECTOR, "[data-message-author-role='assistant']")
        
        if messages:
            last_message = messages[-1]
            response_text = last_message.text
            
            if response_text:
                print(f"  ✅ Ответ получен ({len(response_text)} символов)")
                return response_text
            else:
                print(f"  ⚠️ Ответ пустой")
                return None
        else:
            print(f"  ⚠️ Сообщения не найдены")
            return None
            
    except Exception as e:
        print(f"  ❌ Ошибка при обработке: {e}")
        return None

# ========== ОСНОВНОЙ КОД ==========

# Загружаем Excel
wb, ws = load_excel()
if not wb:
    print("\n❌ Не могу продолжить без Excel файла")
    input("Нажмите ENTER для выхода...")
    exit(1)

# Проверяем невыполненные запросы
pending = get_pending_requests(ws)

if not pending:
    print("✅ Все запросы уже выполнены!")
    input("Нажмите ENTER для выхода...")
    exit(0)

print(f"\n📊 Статистика:")
print(f"   • Невыполненных запросов: {len(pending)}")
print(f"   • Файл: {os.path.abspath(EXCEL_FILE)}")
print("-" * 70)

# Запускаем браузер
print("\n🌐 Запускаю браузер...")

options = uc.ChromeOptions()
options.binary_location = "C:/Program Files/BraveSoftware/Brave-Browser/Application/brave.exe"

profile_dir = os.path.abspath(os.path.join(os.getcwd(), "chatgpt_profile"))
os.makedirs(profile_dir, exist_ok=True)

options.add_argument(f"--user-data-dir={profile_dir}")

driver = None

try:
    driver = uc.Chrome(options=options, version_main=None)
    print("✅ Браузер запущен!")
    
    print("\n📱 Открываю ChatGPT...")
    driver.get("https://chat.openai.com/")
    
    # ШАГ 1: РУЧНОЙ ВХОД
    print("\n" + "=" * 70)
    print("📋 ШАГ 1: АВТОРИЗАЦИЯ (РУЧНОЙ РЕЖИМ)")
    print("=" * 70)
    print("✋ Сейчас ваша очередь:")
    print("   1. Залогиньтесь в ChatGPT (если нужно)")
    print("   2. Введите код из email (если попросит)")
    print("   3. Дождитесь полной загрузки главной страницы")
    print("   4. Убедитесь что вы видите поле ввода внизу")
    print("\n👉 После успешного входа нажмите ENTER в этой консоли")
    print("=" * 70)
    
    input("\n⏸️  Нажмите ENTER когда будете готовы >>> ")
    
    print("\n✅ Отлично! Даю странице еще 5 секунд на загрузку...")
    time.sleep(5)
    
    # ШАГ 2: АВТОМАТИЧЕСКАЯ ОБРАБОТКА
    print("\n" + "=" * 70)
    print("📋 ШАГ 2: АВТОМАТИЧЕСКАЯ ОБРАБОТКА ЗАПРОСОВ")
    print("=" * 70)
    print(f"🔄 Начинаю обрабатывать {len(pending)} запросов из Excel...")
    print("💡 Вы можете наблюдать за процессом в браузере")
    print("=" * 70)
    
    # Обрабатываем каждый запрос
    success_count = 0
    error_count = 0
    
    for idx, item in enumerate(pending, 1):
        row = item['row']
        request = item['request']
        
        print(f"\n{'='*70}")
        print(f"📝 Запрос {idx}/{len(pending)} (строка Excel: {row})")
        print(f"💬 Текст: '{request[:70]}{'...' if len(request) > 70 else ''}'")
        print(f"{'='*70}")
        
        # Помечаем как "В процессе"
        update_status(ws, row, "В процессе")
        save_excel(wb)
        
        # Отправляем в ChatGPT
        response = send_to_chatgpt(driver, request)
        
        if response:
            # Успех
            update_status(ws, row, "Выполнен", response)
            success_count += 1
            print(f"  🎉 Запрос выполнен успешно!")
            print(f"  📄 Первые 150 символов ответа:")
            print(f"  {response[:150]}...")
        else:
            # Ошибка
            update_status(ws, row, "Ошибка", "Не удалось получить ответ")
            error_count += 1
            print(f"  ⚠️ Ошибка при выполнении запроса")
        
        # Сохраняем прогресс после каждого запроса
        save_excel(wb)
        
        # Пауза между запросами (кроме последнего)
        if idx < len(pending):
            print(f"\n  ⏸️  Пауза {DELAY_BETWEEN_REQUESTS} секунд перед следующим запросом...")
            time.sleep(DELAY_BETWEEN_REQUESTS)
    
    # ШАГ 3: РЕЗУЛЬТАТЫ
    print("\n" + "=" * 70)
    print("📋 ШАГ 3: ОБРАБОТКА ЗАВЕРШЕНА!")
    print("=" * 70)
    print(f"✅ Выполнено успешно: {success_count}")
    print(f"❌ Ошибок: {error_count}")
    print(f"📂 Результаты сохранены в: {os.path.abspath(EXCEL_FILE)}")
    print("=" * 70)
    
    # ШАГ 4: РУЧНОЕ ЗАКРЫТИЕ
    print("\n" + "=" * 70)
    print("📋 ШАГ 4: ЗАВЕРШЕНИЕ (РУЧНОЙ РЕЖИМ)")
    print("=" * 70)
    print("✋ Браузер останется открытым")
    print("   • Можете проверить результаты в браузере")
    print("   • Можете посмотреть историю диалога")
    print("   • Можете вручную отправить еще запросы")
    print("\n👉 Когда закончите - нажмите ENTER для закрытия браузера и выхода")
    print("=" * 70)
    
    input("\n⏸️  Нажмите ENTER когда будете готовы закрыть браузер >>> ")
    
    print("\n✅ Закрываю браузер...")

except KeyboardInterrupt:
    print("\n\n⚠️ Программа прервана (Ctrl+C)")
    print("💾 Весь прогресс сохранен в Excel")
    save_excel(wb)

except Exception as e:
    print(f"\n❌ Произошла критическая ошибка: {e}")
    import traceback
    traceback.print_exc()
    save_excel(wb)
    
    print("\n⚠️ Браузер останется открытым для диагностики")
    input("Нажмите ENTER для закрытия...")

finally:
    if driver:
        try:
            driver.quit()
            print("✅ Браузер закрыт")
        except:
            pass
    
    print("\n" + "=" * 70)
    print("✅ Программа завершена")
    print("=" * 70)
    input("\nНажмите ENTER для выхода...")