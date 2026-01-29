"""
Валидация Excel файла и данных
"""
from openpyxl import load_workbook
import os
from config import *

class Validator:
    """Класс для валидации данных"""
    
    def __init__(self):
        self.errors = []
        self.warnings = []
    
    def validate_file_exists(self, filename):
        """Проверяет существование файла"""
        if not os.path.exists(filename):
            self.errors.append(f"Файл {filename} не найден")
            return False
        return True
    
    def validate_excel_structure(self, filename):
        """Проверяет структуру Excel файла"""
        if not self.validate_file_exists(filename):
            return False
        
        try:
            wb = load_workbook(filename)
            ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
            
            # Проверяем минимальное количество строк
            if ws.max_row < 2:
                self.errors.append("Excel файл пустой (нет данных кроме заголовка)")
                return False
            
            # Проверяем заголовки
            expected_headers = {
                COL_REQUEST: "Запрос",
                COL_RESPONSE: "Ответ", 
                COL_STATUS: "Статус",
                COL_DATE: "Дата выполнения",
                COL_ERROR: "Ошибка"
            }
            
            for col, expected_name in expected_headers.items():
                actual_value = ws.cell(1, col).value
                if actual_value != expected_name:
                    self.warnings.append(
                        f"Колонка {col}: ожидался заголовок '{expected_name}', "
                        f"найден '{actual_value}'"
                    )
            
            # Проверяем что есть хотя бы один запрос
            has_requests = False
            empty_rows = []
            
            for row in range(2, ws.max_row + 1):
                request = ws.cell(row, COL_REQUEST).value
                if request and str(request).strip():
                    has_requests = True
                else:
                    empty_rows.append(row)
            
            if not has_requests:
                self.errors.append("Нет ни одного запроса для обработки")
                return False
            
            if empty_rows:
                self.warnings.append(
                    f"Найдены пустые строки: {', '.join(map(str, empty_rows[:5]))}"
                    f"{'...' if len(empty_rows) > 5 else ''}"
                )
            
            wb.close()
            return True
            
        except Exception as e:
            self.errors.append(f"Ошибка при чтении Excel: {e}")
            return False
    
    def validate_browser_path(self, path):
        """Проверяет существование браузера"""
        if not os.path.exists(path):
            self.errors.append(f"Браузер не найден по пути: {path}")
            return False
        return True
    
    def validate_all(self):
        """Выполняет все проверки"""
        print("\n🔍 Валидация конфигурации...")
        print("-" * 70)
        
        valid = True
        
        # Проверяем Excel
        print("📄 Проверяю Excel файл...")
        if not self.validate_excel_structure(EXCEL_FILE):
            valid = False
        else:
            print("   ✅ Excel файл корректен")
        
        # Проверяем браузер
        print("🌐 Проверяю путь к браузеру...")
        if not self.validate_browser_path(BRAVE_PATH):
            valid = False
        else:
            print("   ✅ Браузер найден")
        
        # Проверяем директорию для профиля
        print("📁 Проверяю директорию профиля...")
        try:
            os.makedirs(PROFILE_DIR, exist_ok=True)
            print("   ✅ Директория профиля готова")
        except Exception as e:
            self.errors.append(f"Не могу создать директорию профиля: {e}")
            valid = False
        
        # Выводим результаты
        print("-" * 70)
        
        if self.errors:
            print("\n❌ ОШИБКИ:")
            for i, error in enumerate(self.errors, 1):
                print(f"   {i}. {error}")
        
        if self.warnings:
            print("\n⚠️  ПРЕДУПРЕЖДЕНИЯ:")
            for i, warning in enumerate(self.warnings, 1):
                print(f"   {i}. {warning}")
        
        if valid and not self.warnings:
            print("\n✅ Все проверки пройдены успешно!")
        elif valid:
            print("\n✅ Проверки пройдены (есть предупреждения)")
        else:
            print("\n❌ Валидация не пройдена!")
        
        print("-" * 70)
        
        return valid
    
    def get_errors(self):
        """Возвращает список ошибок"""
        return self.errors
    
    def get_warnings(self):
        """Возвращает список предупреждений"""
        return self.warnings